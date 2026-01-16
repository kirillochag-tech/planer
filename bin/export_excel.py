# bin/export_excel.py
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import QFileDialog, QMessageBox
from PySide6.QtCore import QLocale
from bin import constant as const_
from bin.get_data import Get_Data


def export_full_dashboard(parent, get_data_func, get_special_groups_data,
                          tab_texts, active_tab_index, target_percent):
    """
    Экспортирует все вкладки приложения в один Excel-файл.
    Первой вкладкой будет та, из которой вызван экспорт.
    """
    all_tab_indices = list(range(len(tab_texts)))
    other_tabs = [i for i in all_tab_indices if i != active_tab_index]
    tab_order = [active_tab_index] + other_tabs

    my_date = datetime.now().strftime('%Y-%m-%d %H.%M.%S')
    file_path, _ = QFileDialog.getSaveFileName(
        parent,
        "Сохранить полный дашборд в Excel",
        os.path.expanduser(f"~/Планерка_полный_экспорт_{my_date}.xlsx"),
        "Excel Files (*.xlsx)"
    )
    if not file_path:
        return
    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'

    try:
        wb = Workbook()
        first_sheet = True

        for tab_idx in tab_order:
            tab_name = tab_texts[tab_idx]
            if tab_idx in [0, 4]:
                df = get_data_func(tab_idx)
                if not df.empty:
                    if first_sheet:
                        ws = wb.active
                        ws.title = tab_name
                        first_sheet = False
                    else:
                        ws = wb.create_sheet(title=tab_name)
                    _export_managers_tab(ws, df, tab_name, target_percent)

                df_sp = get_special_groups_data(tab_idx)
                if not df_sp.empty:
                    sp_sheet_name = f"Спецгруппы ({tab_name})"
                    ws_sp = wb.create_sheet(title=sp_sheet_name)
                    _export_special_groups_tab(ws_sp, df_sp,
                                               tab_name, target_percent)

            elif tab_idx in [1, 5]:
                df = get_data_func(tab_idx)
                if not df.empty:
                    if first_sheet:
                        ws = wb.active
                        ws.title = tab_name
                        first_sheet = False
                    else:
                        ws = wb.create_sheet(title=tab_name)
                    _export_brand_managers_tab(ws, df,
                                               tab_name, target_percent)

            elif tab_idx == 2:
                df = get_data_func(tab_idx)
                if not df.empty:
                    if first_sheet:
                        ws = wb.active
                        ws.title = tab_name
                        first_sheet = False
                    else:
                        ws = wb.create_sheet(title=tab_name)
                    _export_brand_managers_farban_tab(ws, df, 
                                                      tab_name, target_percent)

        wb.save(file_path)
        QMessageBox.information(parent, "Экспорт", f"Полный дашборд успешно сохранён:\n{file_path}")

    except Exception as e:
        QMessageBox.critical(parent, "Ошибка", f"Не удалось сохранить файл:\n{e}")


def _get_file_last_modified(file_path: str) -> str:
    """Возвращает форматированную дату и время последнего изменения файла."""
    try:
        mtime = os.path.getmtime(file_path)
        dt = datetime.fromtimestamp(mtime)
        return dt.strftime('%d.%m.%Y %H:%M')
    except (OSError, ValueError):
        return "неизвестно"


def _get_common_styles():
    return {
        'header_fill': PatternFill(start_color="192028", end_color="192028", fill_type="solid"),
        'header_font': Font(color="FFFFFF", bold=True),
        'yellow_fill': PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid"),
        'thin_border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'align_center': Alignment(horizontal="center", vertical="center"),
        'align_left': Alignment(horizontal="left", vertical="center"),
        'color_map': {
            'green': '4CAF50',
            'red': 'F44336',
            'yellow': 'FFEB3B',
            'default': 'BBDEFB',
            'border': '2d3847'
        }
    }


def _format_number(val):
    if pd.isna(val) or val is None:
        return "0"
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        _app_locale = QLocale(QLocale.Russian, QLocale.Russia)
        return _app_locale.toString(float(val), 'f', 0)
    return str(val)


def _add_header_row(ws, tab_name, target_percent, start_col=1):
    """Добавляет строку 'По плану требуется... Данные на...' в начало листа."""
    styles = _get_common_styles()

    file_rel = const_.DICT_TO_TABS.get(tab_name, "")
    file_path = os.path.join("files", file_rel)
    last_modified = _get_file_last_modified(file_path)
    header_text = f"По плану требуется {target_percent} % Данные на: {last_modified}"

    cell = ws.cell(1, start_col, header_text)
    cell.fill = styles['header_fill']
    cell.font = Font(color="FFFFFF", bold=True, size=11)
    cell.alignment = styles['align_center']
    cell.border = styles['thin_border']

    # Определяем, сколько колонок объединить
    max_col = ws.max_column or 10
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=max_col)
    return 2  # данные начнутся со 2-й строки


def _export_managers_tab(ws, df, tab_name, target_percent):
    styles = _get_common_styles()
    data_start_row = _add_header_row(ws, tab_name, target_percent)

    headers = ['Менеджер', 'План (деньги)', 'Факт (деньги)', 'Процент (%)',
               'План (маржа)', 'Факт (маржа)', 'Процент (%)',
               'План (продажи)', 'Факт (продажи)', 'Процент (%)']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(data_start_row, col_num, header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']

    for row_idx, row in df.iterrows():
        if row['manager'] in ('Менеджер', 'Направление'):
            continue
        excel_row = [
            row['manager'],
            _format_number(row['money_plan']),
            _format_number(row['money_fact']),
            _format_number(row['money_percent']),
            _format_number(row['margin_plan']),
            _format_number(row['margin_fact']),
            _format_number(row['margin_percent']),
            _format_number(row['realization_plan']),
            _format_number(row['realization_fact']),
            _format_number(row['realization_percent']),
        ]
        colors = [
            None, 'yellow', row['money_color'], row['money_color'],
            'yellow', row['margin_color'], row['margin_color'],
            'yellow', row['realization_color'], row['realization_color']
        ]
        for col_num, (value, color_key) in enumerate(zip(excel_row, colors), 1):
            cell = ws.cell(row_idx + data_start_row, col_num, value)
            cell.alignment = styles['align_left'] if col_num == 1 else styles['align_center']
            cell.border = styles['thin_border']
            if color_key and color_key in styles['color_map']:
                fill_color = styles['color_map'][color_key]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    _auto_adjust_column_width(ws)


def _export_brand_managers_tab(ws, df, tab_name, target_percent):
    styles = _get_common_styles()
    data_start_row = _add_header_row(ws, tab_name, target_percent)

    headers = ['Менеджер', 'План (все)', 'Факт (все)', 'Процент (%)',
               'Группа', 'План (группа)', 'Факт (группа)', 'Процент (%)']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(data_start_row, col_num, header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']

    current_row = data_start_row + 1
    current_manager = None

    for _, row in df.iterrows():
        if row['manager'] in ('Менеджер', 'Направление'):
            continue
        if row['manager'] != current_manager:
            excel_row = [
                row['manager'],
                _format_number(row['manager_plan']),
                _format_number(row['manager_realization']),
                row['manager_percent'],
                "", "", "", ""
            ]
            for col_num, value in enumerate(excel_row, 1):
                cell = ws.cell(current_row, col_num, value)
                cell.alignment = styles['align_left'] if col_num == 1 else styles['align_center']
                cell.border = styles['thin_border']
                if col_num == 2:
                    cell.fill = styles['yellow_fill']
            current_row += 1
            current_manager = row['manager']

        if row['group']:
            excel_row = [
                "", "", "", "",
                row['group'],
                _format_number(row['group_plan']),
                _format_number(row['group_realization']),
                row['group_percent']
            ]
            colors = [None, None, None, None, None, 'yellow', row['color_cell'], row['color_cell']]
            for col_num, (value, color_key) in enumerate(zip(excel_row, colors), 1):
                cell = ws.cell(current_row, col_num, value)
                cell.alignment = styles['align_center']
                cell.border = styles['thin_border']
                if color_key and color_key in styles['color_map']:
                    fill_color = styles['color_map'][color_key]
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            current_row += 1

    _auto_adjust_column_width(ws)


def _export_brand_managers_farban_tab(ws, df, tab_name, target_percent):
    styles = _get_common_styles()
    data_start_row = _add_header_row(ws, tab_name, target_percent)

    headers = ['Менеджер', 'План (продажи)', 'Факт (продажи)', 'Процент (%)',
               'План (вес)', 'Факт (вес)', 'Процент (%)']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(data_start_row, col_num, header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']

    current_row = data_start_row + 1
    current_manager = None

    for _, row in df.iterrows():
        if row['manager'] in ('Менеджер', 'Направление'):
            continue
        if row['manager'] != current_manager:
            excel_row = [
                row['manager'],
                _format_number(row['manager_plan']),
                _format_number(row['manager_fact']),
                row['manager_percent'],
                _format_number(row['manager_plan_weight']),
                _format_number(row['manager_fact_weight']),
                row['manager_percent_weight']
            ]
            colors = [None, 'yellow', None, None, 'yellow', None, None]
            for col_num, (value, color_key) in enumerate(zip(excel_row, colors), 1):
                cell = ws.cell(current_row, col_num, value)
                cell.alignment = styles['align_left'] if col_num == 1 else styles['align_center']
                cell.border = styles['thin_border']
                if color_key == 'yellow':
                    cell.fill = styles['yellow_fill']
            current_row += 1
            current_manager = row['manager']

        if row['group']:
            excel_row = [
                row['group'],
                _format_number(row['group_plan']),
                _format_number(row['group_fact']),
                row['group_percent'],
                _format_number(row['group_plan_weight']),
                _format_number(row['group_fact_weight']),
                row['group_percent_weight']
            ]
            colors = [
                None,
                'yellow', row['color_cell'], row['color_cell'],
                'yellow', row['color_cell_weight'], row['color_cell_weight']
            ]
            for col_num, (value, color_key) in enumerate(zip(excel_row, colors), 1):
                cell = ws.cell(current_row, col_num, value)
                cell.alignment = styles['align_left'] if col_num == 1 else styles['align_center']
                cell.border = styles['thin_border']
                if color_key and color_key in styles['color_map']:
                    fill_color = styles['color_map'][color_key]
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            current_row += 1

    _auto_adjust_column_width(ws)


def _export_special_groups_tab(ws, df, tab_name, target_percent):
    styles = _get_common_styles()
    data_start_row = _add_header_row(ws, tab_name, target_percent)

    all_groups = sorted([
        g for g in df['special_group'].dropna().unique()
        if g is not None and g != '' and g != 'Все спецгруппы'
    ])

    headers = ['Менеджер']
    for grp in all_groups:
        headers.extend([f'{grp} План', f'{grp} Факт', f'{grp} %'])

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(data_start_row, col_num, header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']

    grouped = df.groupby('manager')
    for row_idx, (manager, group_df) in enumerate(grouped, data_start_row + 1):
        col = 1
        cell = ws.cell(row_idx, col, manager)
        cell.alignment = styles['align_left']
        cell.border = styles['thin_border']
        col += 1

        group_dict = {r['special_group']: r for r in group_df.to_dict('records')}
        for grp in all_groups:
            rec = group_dict.get(grp)
            if rec is not None:
                plan_val = _format_number(rec['special_group_plan'])
                fact_val = _format_number(rec['special_group_fact'])
                pct_val = f"{rec['special_group_percent']:.1f} %" if pd.notna(rec['special_group_percent']) else "0.0 %"
                color_key = rec['special_group_color']

                # План
                plan_cell = ws.cell(row_idx, col, plan_val)
                plan_cell.fill = styles['yellow_fill']
                plan_cell.alignment = styles['align_center']
                plan_cell.border = styles['thin_border']
                col += 1

                # Факт
                fact_cell = ws.cell(row_idx, col, fact_val)
                if color_key in styles['color_map']:
                    hex_color = styles['color_map'][color_key]
                    fact_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                fact_cell.alignment = styles['align_center']
                fact_cell.border = styles['thin_border']
                col += 1

                # %
                pct_cell = ws.cell(row_idx, col, pct_val)
                if color_key in styles['color_map']:
                    hex_color = styles['color_map'][color_key]
                    pct_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                pct_cell.alignment = styles['align_center']
                pct_cell.border = styles['thin_border']
                col += 1
            else:
                for _ in range(3):
                    empty = ws.cell(row_idx, col, "")
                    empty.border = styles['thin_border']
                    empty.alignment = styles['align_center']
                    col += 1

    # Итоги по спецгруппам
    if not df.empty:
        totals_row = len(grouped) + data_start_row + 1
        ws.cell(totals_row, 1, "Итого по спецгруппам").fill = styles['header_fill']
        ws.cell(totals_row, 1).font = styles['header_font']
        ws.cell(totals_row, 1).alignment = styles['align_center']
        ws.cell(totals_row, 1).border = styles['thin_border']

        col = 2
        for grp in all_groups:
            group_data = df[df['special_group'] == grp]
            plan = group_data['special_group_plan'].sum()
            fact = group_data['special_group_fact'].sum()
            percent = (fact / plan * 100) if plan != 0 else 0.0
            color = 'green' if percent >= target_percent else 'red'

            # План
            ws.cell(totals_row, col, _format_number(plan)).fill = styles['yellow_fill']
            ws.cell(totals_row, col).alignment = styles['align_center']
            ws.cell(totals_row, col).border = styles['thin_border']
            col += 1

            # Факт
            fact_cell = ws.cell(totals_row, col, _format_number(fact))
            hex_color = styles['color_map'].get(color, 'FFFFFF')
            fact_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            fact_cell.alignment = styles['align_center']
            fact_cell.border = styles['thin_border']
            col += 1

            # %
            pct_cell = ws.cell(totals_row, col, f"{percent:.1f} %")
            pct_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            pct_cell.alignment = styles['align_center']
            pct_cell.border = styles['thin_border']
            col += 1

    _auto_adjust_column_width(ws)


def _auto_adjust_column_width(ws):
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_len + 2, 60 if idx == 1 else 20)
        ws.column_dimensions[column].width = adjusted_width